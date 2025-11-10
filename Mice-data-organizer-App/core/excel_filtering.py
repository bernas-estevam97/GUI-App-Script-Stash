import os
import pandas as pd
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def filter_excel_by_column(index_file_total_tuple):
    index, file, total_files, folder_output = index_file_total_tuple
    try:
        # Load the Excel file
        df_raw = pd.read_excel(file, sheet_name='Raw Data')

        column_name_start = '/Feature/Tail/Tip_X'
        valid_indices_beg = df_raw[df_raw[column_name_start].notna()].index

        if valid_indices_beg.empty:
            print(f"[WARNING] File '{file}': No valid data found in column '{column_name_start}'. Skipping...")
            return
        
        start_index = valid_indices_beg[0]
        last_index = df_raw['/Feature/Head/Nose_X'].last_valid_index()

        # Filter the dataframe in another sheet
        df_calc = pd.read_excel(file, sheet_name='Calculated Data')
        filtered_df = df_calc.iloc[start_index:last_index+1]

        if 'Time' not in filtered_df.columns or filtered_df['Time'].dropna().empty:
            print(f"[WARNING] File '{file}': 'Time' column missing or empty. Skipping...")
            return

        # Calculate time range
        start_time = filtered_df['Time'].dropna().iloc[0]
        stop_time = filtered_df['Time'].dropna().iloc[-1]

        # Append a row with column averages
        averages = filtered_df.mean(numeric_only=True)
        averages_row = pd.DataFrame([averages], columns=filtered_df.columns)

        # Make sure first column is object type before inserting time diff
        averages_row = averages_row.astype({averages_row.columns[0]: 'object'})
        averages_row.iloc[0, 0] = stop_time - start_time

        # Append both label row and averages row
        filtered_df = pd.concat([filtered_df, averages_row], ignore_index=True)

        # Determine output file path
        output_dir = folder_output if folder_output and os.path.isdir(folder_output) else os.path.dirname(file)
        os.makedirs(output_dir, exist_ok=True)

        output_file = os.path.join(
            output_dir,
            os.path.splitext(os.path.basename(file))[0] + "_filtered.xlsx"
        )

        # Save to Excel
        filtered_df.to_excel(output_file, sheet_name='Calculated Data', index=False)

        # Add merged row
        wb = load_workbook(output_file)
        ws = wb['Calculated Data']

        mean_row_index = ws.max_row
        ws.insert_rows(mean_row_index)
        num_cols = ws.max_column
        start_col = get_column_letter(1)
        end_col = get_column_letter(num_cols)
        ws.merge_cells(f"{start_col}{mean_row_index}:{end_col}{mean_row_index}")
        ws[f"{start_col}{mean_row_index}"] = "MEANS LINE UNDER"

        wb.save(output_file)
        wb.close()

        return f"{os.path.basename(file)} → {os.path.basename(output_file)}"

    except Exception as e:
        print(f"[ERROR] File '{file}': {e}. Skipping...")


def run(folder_input, folder_output=None, log_fn=print):
    """Run the Excel filtering process."""
    if not folder_input:
        log_fn("❌ No input path provided.")
        return
    if not os.path.isdir(folder_input):
        log_fn(f"❌ Invalid path: {folder_input}")
        return

    # Default output folder = same as input
    if not folder_output or not os.path.isdir(folder_output):
        folder_output = folder_input
        log_fn("📁 No output folder selected. Using input folder as output.")

    log_fn(f"📂 Input folder: {os.path.normpath(folder_input)}")
    log_fn(f"📦 Output folder: {os.path.normpath(folder_output)}")

    # Collect Excel files
    file_paths = [
        os.path.join(folder_input, f)
        for f in os.listdir(folder_input)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    ]
    if not file_paths:
        log_fn(f"⚠️ No Excel files found in:\n   {folder_input}")
        return

    total_files = len(file_paths)
    log_fn(f"📊 Found {total_files} Excel file(s) to process")

    num_processes = os.cpu_count()
    log_fn(f"⚙️ Starting parallel processing with {num_processes} CPU cores...")

    start = time.perf_counter()

    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        futures = [
            executor.submit(filter_excel_by_column, (i, f, total_files, folder_output))
            for i, f in enumerate(file_paths)
        ]
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    log_fn(f"✅ {result}")
            except Exception as exc:
                log_fn(f"❌ Error while processing file: {exc}")

    finish = time.perf_counter()
    log_fn(f"🎉 All files processed in {round(finish - start, 2)} seconds.")



def main():
    folder_input = input('Which folder has your .xlsx files? ').strip()
    if not folder_input:
        print("❌ You didn't input any path.")
        return
    if not os.path.isdir(folder_input):
        print(f"❌ Invalid path: {folder_input}")
        return

    print(f"📁 Folder selected: {folder_input}")

    # Step 2: Get .xlsx files
    file_paths = [
        os.path.join(folder_input, f)
        for f in os.listdir(folder_input)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')  # ignore temp Excel files
    ]

    if not file_paths:
        print(f"⚠️ No .xlsx files found in: {folder_input}")
        return

    total_files = len(file_paths)
    indexed_files = [(index, file_path, total_files) for index, file_path in enumerate(file_paths)]

    # Step 3: Process files in parallel
    num_processes = os.cpu_count()
    print(f"⚙️ Starting processing on {num_processes} CPU cores...")

    start = time.perf_counter()

    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        futures = [executor.submit(filter_excel_by_column, indexed_file) for indexed_file in indexed_files]

        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    print(result)
            except Exception as exc:
                print(f"❌ Error: {exc}")

    finish = time.perf_counter()
    print(f"✅ All files processed in {round(finish - start, 2)} seconds.")
                
if __name__ == '__main__':
    # Keep CLI usability for testing
    main()
    
    
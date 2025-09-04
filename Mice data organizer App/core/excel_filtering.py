import os
import pandas as pd
import time
from multiprocessing import Pool
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def filter_excel_by_column(index_file_total_tuple):
    index, file, total_files = index_file_total_tuple
    try:
        # Load the Excel file
        df_raw = pd.read_excel(file, sheet_name='Raw Data')

        column_name_start = '/Feature/Tail/Tip_X'
        # Find the first index where the column has a non-null value
        valid_indices_beg = df_raw[df_raw[column_name_start].notna()].index

        if valid_indices_beg.empty:
            print(f"[WARNING] File '{file}': No valid data found in column '{column_name_start}'. Skipping...")
            return
        
        start_index = valid_indices_beg[0]
        last_index = df_raw['/Feature/Head/Nose_X'].last_valid_index()

        # Filter the dataframe in another sheet
        df_calc = pd.read_excel(file, sheet_name='Calculated Data')
        filtered_df = df_calc.iloc[start_index:last_index+1]

        # Ensure the 'Time' column is present and valid
        if 'Time' not in filtered_df.columns or filtered_df['Time'].dropna().empty:
            print(f"[WARNING] File '{file}': 'Time' column missing or empty in filtered data. Skipping...")
            return

        # Calculate time range
        start_time = filtered_df['Time'].dropna().iloc[0]
        stop_time = filtered_df['Time'].dropna().iloc[-1]

        # Append a row with column averages
        averages = filtered_df.mean(numeric_only=True)
        averages_row = pd.DataFrame([averages], columns=filtered_df.columns)

        # Ensure the first column is string type before assignment
        averages_row = averages_row.astype({averages_row.columns[0]: 'object'})
        averages_row.iloc[0, 0] = stop_time - start_time

        # Append both label row and averages row
        filtered_df = pd.concat([filtered_df, averages_row], ignore_index=True)

        # Save to a new Excel file
        output_file = os.path.splitext(file)[0]+'_filtered'+os.path.splitext(file)[1]
        filtered_df.to_excel(output_file, sheet_name='Calculated Data', index=False)

        # Add merged cell row using openpyxl
        wb = load_workbook(output_file)
        ws = wb['Calculated Data']

        mean_row_index = ws.max_row
        ws.insert_rows(mean_row_index)
        num_cols = ws.max_column
        start_col = get_column_letter(1)
        end_col = get_column_letter(num_cols)
        ws.merge_cells(f"{start_col}{mean_row_index}:{end_col}{mean_row_index}")
        ws[f"{start_col}{mean_row_index}"] = "MEANS LINE UNDER"

        wb.save(output_file)
        return f"Processed file {index + 1} of {total_files}: {os.path.basename(file)} ---> {os.path.basename(output_file)}"

    except Exception as e:
        # Catch and log any unexpected error
        print(f"[ERROR] File '{file}': {e}. Skipping...")

# Example usage



def run(folder_input, log_fn=print):
    if not folder_input:
        log_fn("❌ No input path provided.")
        return
    if not os.path.isdir(folder_input):
        log_fn(f"❌ Invalid path: {folder_input}")
        return

    log_fn(f"📂 Selected folder: {os.path.basename(os.path.normpath(folder_input))}")

    file_paths = [
        os.path.join(folder_input, f)
        for f in os.listdir(folder_input)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    ]

    if not file_paths:
        log_fn(f"⚠️ No Excel files found in:\n   {folder_input}")
        return

    total_files = len(file_paths)
    log_fn(f"📊 Found {total_files} Excel file(s) to process")

    num_processes = os.cpu_count()
    log_fn(f"⚙️ Starting parallel processing with {num_processes} CPU cores...")

    start = time.perf_counter()

    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        futures = [executor.submit(filter_excel_by_column, (i, f, total_files))
                   for i, f in enumerate(file_paths)]
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    log_fn(f"✅ Finished: {os.path.basename(result)}")
            except Exception as exc:
                log_fn(f"❌ Error while processing file: {exc}")

    finish = time.perf_counter()
    log_fn(f"🎉 All files processed in {round(finish - start, 2)} seconds.")


def main():
    folder_input = input('Which folder has your .xlsx files? ').strip()
    if not folder_input:
        print("❌ You didn't input any path.")
        return
    if not os.path.isdir(folder_input):
        print(f"❌ Invalid path: {folder_input}")
        return

    print(f"📁 Folder selected: {folder_input}")

    # Step 2: Get .xlsx files
    file_paths = [
        os.path.join(folder_input, f)
        for f in os.listdir(folder_input)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')  # ignore temp Excel files
    ]

    if not file_paths:
        print(f"⚠️ No .xlsx files found in: {folder_input}")
        return

    total_files = len(file_paths)
    indexed_files = [(index, file_path, total_files) for index, file_path in enumerate(file_paths)]

    # Step 3: Process files in parallel
    num_processes = os.cpu_count()
    print(f"⚙️ Starting processing on {num_processes} CPU cores...")

    start = time.perf_counter()

    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        futures = [executor.submit(filter_excel_by_column, indexed_file) for indexed_file in indexed_files]

        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    print(result)
            except Exception as exc:
                print(f"❌ Error: {exc}")

    finish = time.perf_counter()
    print(f"✅ All files processed in {round(finish - start, 2)} seconds.")
                
if __name__ == '__main__':
    # Keep CLI usability for testing
    main()
    
    
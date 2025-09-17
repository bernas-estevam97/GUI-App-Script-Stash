


import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
from collections import defaultdict
import numpy as np

# All possible target labels in Excel
RAW_STRINGS = [
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:"
]

# Map specific labels to unified output names
RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:"
}


data_ids = {
    "3879": ("TG", "Female"),
    "3880": ("TG", "Female"),
    "3881": ("TG", "Male"),
    "3882": ("TG", "Male"),
    "3883": ("WT", "Male"),
    "3884": ("TG", "Male"),
    "3885": ("TG", "Female"),
    "3886": ("TG", "Male"),
    "3968": ("WT", "Female"),
    "3969": ("TG", "Female"),
    "3970": ("WT", "Male"),
    "3971": ("TG", "Male"),
    "3972": ("WT", "Male"),
    "3994": ("TG", "Female"),
    "3995": ("TG", "Male"),
    "3996": ("TG", "Male"),
    "3997": ("TG", "Male"),
    "3998": ("TG", "Male"),
    "3999": ("TG", "Male"),
    "2200_1L":   ("WT", "Male"),
    "2200_1R":   ("WT", "Male"),
    "2199_1R": ("WT", "Female"),
    "2221_1L": ("WT", "Female"),
    "2221_1R": ("WT", "Female"),
    "2221_1RL":  ("WT", "Female"),
    "2222_1L": ("WT", "Male"),
    "2222_1R": ("WT", "Male"),
    "2222_1RL":("WT", "Male"),
    "2273_1L": ("TG", "Female"),
    "2273_1R": ("TG", "Female"),
    "2273_1RL":("TG", "Female"),
    "2273_2R":   ("TG", "Female"),
    "2280_1R": ("WT", "Female"),
    "2280_1B":  ("WT", "Female"),
    "2512_1L": ("WT", "Male"),
    "2512_1R": ("WT", "Male"),
    "2517_1R ": ("WT", "Male"),
    "2886": ("WT", "Male"),
    "2887": ("WT", "Male"),
    "3000": ("WT", "Female"),
    "3043": ("TG", "Female"),
    "3051": ("WT", "Male"),
    "3054": ("WT", "Male"),
    "3055": ("TG", "Male"),
    "3151": ("WT", "Female"),
    "3153": ("TG", "Male"),
    "3156": ("TG", "Male"),
    "3220": ("TG", "Female"),
    "3223": ("TG", "Male"),
    "3272": ("WT", "Female"),
    "3274": ("WT", "Female"),
    "3275": ("TG", "Female"),
    "3276": ("TG", "Female"),
    "3278": ("TG", "Male"),
    "3607": ("TG", "Female")
}









def find_table_data(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    result_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Table ID"):
                    
                    # Extract clean Table ID
                    table_id = cell.value.strip().replace("Table ID:", "").strip()
                    table_values = defaultdict(list)
                    table_values["Table ID"].append(table_id)

                    col_index = cell.column
                    row_index = cell.row

                    for r in range(row_index + 1, ws.max_row + 1):
                        label_cell = ws.cell(row=r, column=col_index)
                        value_cell = ws.cell(row=r, column=col_index + 1)

                        label = label_cell.value
                        value = value_cell.value

                        if label in RAW_STRINGS and isinstance(value, (int, float)):
                            output_label = RENAME_MAP.get(label, label)
                            table_values[output_label].append(value)

                    # Now calculate means for each label
                    table_data = {}
                    for key, values in table_values.items():
                        if key == "Table ID":
                            table_data[key] = values[0]  # Only one Table ID
                        else:
                            table_data[key] = float(np.mean(values)) if values else None

                    result_data.append(table_data)

    return result_data


#Info

# table_data = {
#     "Table ID": "2199_1R_T1_4W",
#     "Stride length left front average in cm:": 7.416,
#     "Stride Width Front average in cm:": 1.45,
#     ...
# }

# This is a single row of data, where:
# keys = column names (like headers in Excel)
# values = values under each column



# result_data = [
#     table_data_1,
#     table_data_2,
#     ...
# ]

#Each key (from the dictionaries) becomes a column.
#Each dictionary becomes a row.
#The dictionary’s key-value pairs are aligned horizontally in Excel or when printed.




# def save_to_excel(data, base_output_dir):
#     # Better formatted timestamp
#     timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     output_file = os.path.join(base_output_dir, f"MFrame_clean_output_{timestamp}.xlsx")
#     df = pd.DataFrame(data)
#     df.to_excel(output_file, index=False)
#     return output_file


# 🚀 This is the entrypoint callable from GUI
def run(input_file):
    from openpyxl.styles import PatternFill, Font

    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    data = find_table_data(input_file)

    if not data:
        raise ValueError("No 'Table ID' entries found in the provided file.")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(os.path.dirname(input_file), f"MFrame_clean_output_{timestamp}.xlsx")

    # Step 1: Write with pandas
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)

    # Step 2: Style with openpyxl
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Define a font for contrast
    white_font = Font(color="FFFFFFFF")

    # Helper function: get fill color per (genotype, sex)
    def get_fill(genotype, sex):
        if (genotype, sex) == ("TG", "Male"):
            return PatternFill(start_color="FF003366", end_color="FF003366", fill_type="solid"), True
        elif (genotype, sex) == ("WT", "Male"):
            return PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid"), False
        elif (genotype, sex) == ("TG", "Female"):
            return PatternFill(start_color="FF8B0000", end_color="FF8B0000", fill_type="solid"), True
        elif (genotype, sex) == ("WT", "Female"):
            return PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid"), False
        else:
            return None, False

    # Get "Table ID" column index
    headers = [cell.value for cell in ws[1]]
    try:
        table_id_col_index = headers.index("Table ID") + 1  # openpyxl is 1-based
    except ValueError:
        raise ValueError("Table ID column not found in Excel output.")

    # Apply row coloring
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        table_id_cell = row[table_id_col_index - 1]
        table_id = str(table_id_cell.value).strip()
        # Match key by substring
        matched_key = next((k for k in data_ids if k in table_id), None)

        if matched_key:
            
            genotype, sex = data_ids[matched_key]
            fill, use_white_font = get_fill(genotype, sex)
            if fill:
                for cell in row:
                    cell.fill = PatternFill(
                        start_color=fill.start_color,
                        end_color=fill.end_color,
                        fill_type="solid"
                    )
                    if use_white_font:
                        cell.font = white_font
    # Save with styling
    wb.save(output_file)
    return output_file



# Keep CLI usability for testing
if __name__ == "__main__":
    input_file = input("Enter the path to the Excel file: ").strip()
    print(f"Input file: {input_file}")
    try:
        output_path = run(input_file)
        print(f"✅ Extracted data has been written to: {output_path}")
    except Exception as e:
        print(f"❌ Error: {e}")
